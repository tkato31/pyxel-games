import unicodedata
import re
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# PyInstallerでexe化したときは sys.executable のフォルダ、
# スクリプト実行時は __file__ のフォルダに設定ファイルを置く
if getattr(sys, "frozen", False):
    _APP_DIR = os.path.dirname(sys.executable)
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_APP_DIR, "settings.json")

DEFAULT_SETTINGS = {
    "fixed_items": ["オプテージ利用料金", "ENEOS-SS", "ﾂｳｺｳﾘﾖｳｷﾝ"],
    "history_file": ""
}


def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return DEFAULT_SETTINGS.copy()


def normalize(text: str) -> str:
    """全角→半角、大文字→小文字に統一して比較用文字列を返す"""
    return unicodedata.normalize("NFKC", text).lower()


def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


# ------------------------------------------------------------------ スタイル定数
def _make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FILL    = PatternFill("solid", fgColor="E60026")
HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
FIXED_FILL  = PatternFill("solid", fgColor="FFCCCC")
MANUAL_FILL = PatternFill("solid", fgColor="FFF9C4")
TOTAL_FONT  = Font(bold=True, size=11, color="C00000")
INFO_FONT   = Font(color="AAAAAA", size=9)


class RakutenApp:
    def __init__(self, root):
        self.root = root
        self.root.title("楽天カード　奥さん請求まとめ")
        self.root.geometry("1000x660")
        self.root.resizable(True, True)
        self.settings = load_settings()
        self.rows = []
        self.checked = []
        self.csv_path = ""
        self.create_widgets()

    # ------------------------------------------------------------------ UI構築
    def create_widgets(self):
        # ヘッダー
        header = tk.Frame(self.root, bg="#e60026", pady=6, padx=12)
        header.pack(fill=tk.X)
        tk.Label(header, text="楽天カード　奥さん請求まとめ",
                 bg="#e60026", fg="white", font=("", 14, "bold")).pack(side=tk.LEFT)

        # ツールバー
        toolbar = tk.Frame(self.root, pady=6, padx=12, bg="#f5f5f5",
                           relief=tk.RIDGE, bd=1)
        toolbar.pack(fill=tk.X)

        tk.Button(toolbar, text="📂 CSVを開く", command=self.load_csv,
                  bg="#e60026", fg="white", font=("", 10, "bold"),
                  padx=10, pady=3, relief=tk.FLAT).pack(side=tk.LEFT)

        tk.Button(toolbar, text="📂 複数CSV一括処理", command=self.batch_process,
                  bg="#a00019", fg="white", font=("", 10, "bold"),
                  padx=10, pady=3, relief=tk.FLAT).pack(side=tk.LEFT, padx=6)

        self.file_label = tk.Label(toolbar, text="ファイルが選択されていません",
                                   fg="gray", bg="#f5f5f5", font=("", 9))
        self.file_label.pack(side=tk.LEFT, padx=12)

        tk.Button(toolbar, text="⚙ 設定", command=self.open_settings,
                  font=("", 9), padx=8, pady=3).pack(side=tk.RIGHT)

        # 凡例
        legend = tk.Frame(toolbar, bg="#f5f5f5")
        legend.pack(side=tk.RIGHT, padx=16)
        tk.Label(legend, text="■", fg="#ffcccc", bg="#f5f5f5", font=("", 13)).pack(side=tk.LEFT)
        tk.Label(legend, text="固定  ", bg="#f5f5f5", font=("", 9)).pack(side=tk.LEFT)
        tk.Label(legend, text="■", fg="#fff9c4", bg="#f5f5f5", font=("", 13)).pack(side=tk.LEFT)
        tk.Label(legend, text="手動選択  ", bg="#f5f5f5", font=("", 9)).pack(side=tk.LEFT)

        # テーブル
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)

        cols = ("date", "store", "user", "method", "amount", "charge")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings")

        self.tree.heading("date",   text="利用日")
        self.tree.heading("store",  text="利用店名・商品名")
        self.tree.heading("user",   text="利用者")
        self.tree.heading("method", text="支払方法")
        self.tree.heading("amount", text="利用金額")
        self.tree.heading("charge", text="請求")

        self.tree.column("date",   width=95,  anchor="center", stretch=False)
        self.tree.column("store",  width=320, anchor="w")
        self.tree.column("user",   width=70,  anchor="center", stretch=False)
        self.tree.column("method", width=90,  anchor="center", stretch=False)
        self.tree.column("amount", width=110, anchor="e",      stretch=False)
        self.tree.column("charge", width=60,  anchor="center", stretch=False)

        style = ttk.Style()
        style.configure("Treeview", rowheight=26, font=("", 10))
        style.configure("Treeview.Heading", font=("", 10, "bold"))

        self.tree.tag_configure("fixed",   background="#ffcccc")
        self.tree.tag_configure("checked", background="#fff9c4")
        self.tree.tag_configure("info",    foreground="#aaaaaa")

        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL,   command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<ButtonRelease-1>", self.on_click)

        # フッター上段：当月請求合計 ＋ Excel出力
        footer_top = tk.Frame(self.root, pady=6, padx=12)
        footer_top.pack(fill=tk.X)

        self.total_label = tk.Label(footer_top, text="請求合計：¥0",
                                    font=("", 13, "bold"), fg="#e60026")
        self.total_label.pack(side=tk.LEFT)

        tk.Button(footer_top, text="📊 Excelに出力", command=self.export_excel,
                  bg="#217346", fg="white", font=("", 11, "bold"),
                  padx=14, pady=4, relief=tk.FLAT).pack(side=tk.RIGHT)

        # フッター下段：残高サマリー ＋ 入金を記録
        footer_bot = tk.Frame(self.root, pady=6, padx=12, bg="#f0f4ff",
                              relief=tk.RIDGE, bd=1)
        footer_bot.pack(fill=tk.X)

        tk.Button(footer_bot, text="💰 入金を記録", command=self.record_payment,
                  bg="#1f4e79", fg="white", font=("", 10, "bold"),
                  padx=10, pady=3, relief=tk.FLAT).pack(side=tk.RIGHT, padx=4)

        self.lbl_invoiced   = tk.Label(footer_bot, text="累計請求：¥―",
                                       font=("", 10), bg="#f0f4ff", fg="#333333")
        self.lbl_received   = tk.Label(footer_bot, text="累計入金：¥―",
                                       font=("", 10), bg="#f0f4ff", fg="#333333")
        self.lbl_balance    = tk.Label(footer_bot, text="未収残高：¥―",
                                       font=("", 11, "bold"), bg="#f0f4ff", fg="#1f4e79")

        self.lbl_invoiced.pack(side=tk.LEFT, padx=(0, 20))
        self.lbl_received.pack(side=tk.LEFT, padx=(0, 20))
        self.lbl_balance.pack(side=tk.LEFT)

        # 起動時に残高を読み込む
        self.root.after(100, self.refresh_balance)

    # ------------------------------------------------------------------ CSV読込
    def load_csv(self):
        filepath = filedialog.askopenfilename(
            title="楽天カードCSVを選択",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not filepath:
            return
        try:
            df = self._read_csv(filepath)
            self.csv_path = filepath
            self.file_label.config(text=os.path.basename(filepath), fg="black")
            rows, checked = self._build_rows(df)
            self.rows    = rows
            self.checked = checked
            self._populate_tree(rows, checked)
        except Exception as e:
            messagebox.showerror("エラー", f"CSVの読み込みに失敗しました。\n{e}")

    def _read_csv(self, filepath):
        for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
            try:
                return pd.read_csv(filepath, encoding=enc)
            except (UnicodeDecodeError, LookupError):
                continue
        raise ValueError("文字コードを判別できませんでした。")

    # ------------------------------------------------------------------ 行データ構築（共通）
    def is_fixed(self, store: str) -> bool:
        store_n = normalize(store)
        return any(normalize(kw) in store_n for kw in self.settings["fixed_items"])

    def safe(self, val) -> str:
        s = str(val).strip()
        return "" if s == "nan" else s

    def _build_rows(self, df):
        """DataFrameから (rows, checked) を生成する"""
        rows    = []
        checked = []
        for _, row in df.iterrows():
            date   = self.safe(row.get("利用日", ""))
            store  = self.safe(row.get("利用店名・商品名", ""))
            user   = self.safe(row.get("利用者", ""))
            method = self.safe(row.get("支払方法", ""))
            try:
                amount = int(float(str(row.get("利用金額", "0")).replace(",", "").strip()))
            except (ValueError, TypeError):
                amount = 0

            is_info  = (date == "")
            is_fixed = self.is_fixed(store) and not is_info

            rows.append({
                "date": date, "store": store, "user": user,
                "method": method, "amount": amount,
                "is_info": is_info, "is_fixed": is_fixed,
            })
            checked.append(is_fixed)
        return rows, checked

    def _populate_tree(self, rows, checked):
        """テーブルを描画する"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, row_data in enumerate(rows):
            amount_str = f"¥{row_data['amount']:,}" if row_data["amount"] else ""
            check      = "○" if checked[i] else ""
            is_info    = row_data["is_info"]
            is_fixed   = row_data["is_fixed"]
            tag        = "info" if is_info else ("fixed" if is_fixed else "")
            self.tree.insert("", tk.END,
                             values=(row_data["date"], row_data["store"],
                                     row_data["user"], row_data["method"],
                                     amount_str, check),
                             tags=(tag,))
        self.update_total()

    # ------------------------------------------------------------------ クリック処理
    def on_click(self, event):
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        idx      = list(self.tree.get_children()).index(item_id)
        row_data = self.rows[idx]
        if row_data["is_info"]:
            return

        self.checked[idx] = not self.checked[idx]
        is_checked = self.checked[idx]
        store = row_data["store"]

        # 固定項目設定へ反映（手動チェック→追加、手動解除→削除）
        if not row_data["is_fixed"]:
            fixed_n = [normalize(kw) for kw in self.settings["fixed_items"]]
            if is_checked and normalize(store) not in fixed_n:
                self.settings["fixed_items"].append(store)
                save_settings(self.settings)
            elif not is_checked and normalize(store) in fixed_n:
                idx_kw = fixed_n.index(normalize(store))
                self.settings["fixed_items"].pop(idx_kw)
                save_settings(self.settings)

        values    = list(self.tree.item(item_id, "values"))
        values[5] = "○" if is_checked else ""

        if row_data["is_fixed"]:
            tag = "fixed"
        elif is_checked:
            tag = "checked"
        else:
            tag = ""

        self.tree.item(item_id, values=values, tags=(tag,))
        self.update_total()

    def update_total(self):
        total = sum(
            r["amount"] for i, r in enumerate(self.rows)
            if self.checked[i] and not r["is_info"]
        )
        self.total_label.config(text=f"請求合計：¥{total:,}")

    # ------------------------------------------------------------------ Excel書き込み（共通）
    def _write_excel_sheet(self, ws, rows, checked):
        """wsにデータを書き込み (total_all, total_wife) を返す"""
        border = _make_border()

        headers = ["利用日", "利用店名・商品名", "利用者", "支払方法", "利用金額", "請求"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = border

        total_all  = 0
        total_wife = 0

        for i, row_data in enumerate(rows):
            r        = i + 2
            chk      = checked[i]
            is_fixed = row_data["is_fixed"]
            is_info  = row_data["is_info"]
            amount   = row_data["amount"] if not is_info else None

            vals = [row_data["date"], row_data["store"],
                    row_data["user"], row_data["method"],
                    amount, "○" if chk else ""]

            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = border
                if is_info:
                    c.font = INFO_FONT
                else:
                    c.font = Font(size=10)
                    if is_fixed:
                        c.fill = FIXED_FILL
                    elif chk:
                        c.fill = MANUAL_FILL
                if col in (1, 3, 4, 6):
                    c.alignment = Alignment(horizontal="center")
                elif col == 5 and amount is not None:
                    c.alignment  = Alignment(horizontal="right")
                    c.number_format = "#,##0"
                else:
                    c.alignment = Alignment(horizontal="left")

            if amount:
                total_all += amount
            if chk and amount:
                total_wife += amount

        # 合計行
        total_row = len(rows) + 3
        lc = ws.cell(row=total_row, column=4, value="請求合計")
        lc.font = Font(bold=True, size=11)
        lc.alignment = Alignment(horizontal="right")

        tc = ws.cell(row=total_row, column=5, value=total_wife)
        tc.font = TOTAL_FONT; tc.number_format = "#,##0"
        tc.alignment = Alignment(horizontal="right")

        ws.cell(row=total_row, column=6, value="円").font = Font(bold=True, size=11)

        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 8
        ws.freeze_panes = "A2"

        return total_all, total_wife

    # ------------------------------------------------------------------ Excel出力（単体）
    def export_excel(self):
        if not self.rows:
            messagebox.showwarning("警告", "先にCSVを読み込んでください。")
            return

        default_name = os.path.splitext(os.path.basename(self.csv_path))[0] if self.csv_path else "明細"
        filepath = filedialog.asksaveasfilename(
            title="Excelファイルを保存",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "楽天カード明細"
            total_all, total_wife = self._write_excel_sheet(ws, self.rows, self.checked)
            wb.save(filepath)
            self._update_history(self.csv_path, total_all, total_wife)
            self.refresh_balance()
            messagebox.showinfo("完了", f"Excelを保存しました。\n\n請求合計：¥{total_wife:,}")
        except Exception as e:
            messagebox.showerror("エラー", f"Excel出力に失敗しました。\n{e}")

    # ------------------------------------------------------------------ 一括処理
    def batch_process(self):
        filepaths = filedialog.askopenfilenames(
            title="楽天カードCSVを複数選択（Ctrl/Shiftで複数選択）",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not filepaths:
            return

        # 履歴ファイルが未設定なら先に選ぶ
        if not self.settings.get("history_file"):
            if not self._choose_history_file():
                return

        overwrite_decision = None  # None=都度確認、True=全て上書き、False=全てスキップ
        results = []
        errors  = []

        for filepath in filepaths:
            try:
                df = self._read_csv(filepath)
                rows, checked = self._build_rows(df)
                total_all, total_wife = self._calc_totals(rows, checked)

                # Excel出力（CSVと同じフォルダに同名.xlsx）
                excel_path = os.path.splitext(filepath)[0] + ".xlsx"
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "楽天カード明細"
                self._write_excel_sheet(ws, rows, checked)
                wb.save(excel_path)

                # 履歴更新
                overwrite_decision = self._update_history(
                    filepath, total_all, total_wife, overwrite_decision
                )

                results.append({
                    "name":       os.path.basename(filepath),
                    "total_all":  total_all,
                    "total_wife": total_wife,
                })
            except Exception as e:
                errors.append(f"{os.path.basename(filepath)}: {e}")

        # 結果サマリー
        msg = f"一括処理完了：{len(results)} 件\n\n"
        for r in results:
            msg += (f"・{r['name']}\n"
                    f"  全請求：¥{r['total_all']:,} ／ 奥さん：¥{r['total_wife']:,}\n")
        if errors:
            msg += f"\nエラー：{len(errors)} 件\n" + "\n".join(errors)
        self.refresh_balance()
        messagebox.showinfo("一括処理完了", msg)

    def _calc_totals(self, rows, checked):
        total_all  = sum(r["amount"] for r in rows if not r["is_info"] and r["amount"])
        total_wife = sum(r["amount"] for i, r in enumerate(rows)
                         if checked[i] and not r["is_info"] and r["amount"])
        return total_all, total_wife

    # ------------------------------------------------------------------ 履歴更新
    def _get_yyyymm(self, filepath):
        m = re.search(r'(\d{4})(\d{2})', os.path.basename(filepath))
        if m:
            return int(m.group(1)), int(m.group(2))
        return None

    def _choose_history_file(self):
        """履歴ファイルの保存先を選ぶ。選択されたらTrue、キャンセルはFalse"""
        path = filedialog.asksaveasfilename(
            title="履歴ファイルの保存場所を選択",
            initialfile="seikyu_history",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return False
        self.settings["history_file"] = path
        save_settings(self.settings)
        return True

    def _update_history(self, filepath, total_all: int, total_wife: int,
                        overwrite_decision=None):
        """
        履歴Excelに当月データを追記 or 上書き。
        overwrite_decision: None=都度確認、True=全て上書き、False=全てスキップ
        戻り値: 更新後の overwrite_decision
        """
        if not self.settings.get("history_file"):
            if not self._choose_history_file():
                return overwrite_decision

        history_path = self.settings["history_file"]

        ym = self._get_yyyymm(filepath)
        if not ym:
            messagebox.showwarning("警告", "CSVファイル名から年月を取得できませんでした。\n履歴は更新されませんでした。")
            return overwrite_decision

        year, month = ym
        label = f"{year}年{month:02d}月"
        diff  = total_all - total_wife

        # スタイル
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        num_fmt  = "#,##0"
        border   = _make_border()
        center   = Alignment(horizontal="center")
        right    = Alignment(horizontal="right")

        def apply_row_style(ws, row_idx):
            fill = PatternFill("solid", fgColor="EBF3FB") if row_idx % 2 == 0 else None
            for c in range(1, 5):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = border
                cell.font   = Font(size=10)
                if fill:
                    cell.fill = fill
                cell.alignment = center if c == 1 else right
                if c > 1:
                    cell.number_format = num_fmt

        # ファイル読込 or 新規作成
        if os.path.exists(history_path):
            wb = openpyxl.load_workbook(history_path)
            ws = wb.active
            self._ensure_payment_sheet(wb)  # Sheet2がなければ追加
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "請求履歴"
            headers = ["年月", "全請求額", "奥さん請求額", "差額"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = center; c.border = border
            ws.column_dimensions["A"].width = 14
            for col in ("B", "C", "D"):
                ws.column_dimensions[col].width = 16
            ws.freeze_panes = "A2"
            # Sheet2「入金履歴」を同時に作成
            self._ensure_payment_sheet(wb)

        # 既存行を検索
        existing_row = None
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value == label:
                existing_row = row[0].row
                break

        if existing_row:
            # 上書き判定
            if overwrite_decision is None:
                if messagebox.askyesno("確認", f"「{label}」のデータが既にあります。上書きしますか？"):
                    overwrite_decision = True
                else:
                    return overwrite_decision  # スキップ（decisionはNoneのまま）
            elif overwrite_decision is False:
                return overwrite_decision  # バッチ：全スキップ

            ws.cell(row=existing_row, column=2, value=total_all)
            ws.cell(row=existing_row, column=3, value=total_wife)
            ws.cell(row=existing_row, column=4, value=diff)
            apply_row_style(ws, existing_row)
        else:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=label)
            ws.cell(row=next_row, column=2, value=total_all)
            ws.cell(row=next_row, column=3, value=total_wife)
            ws.cell(row=next_row, column=4, value=diff)
            apply_row_style(ws, next_row)

        # 年月順ソート
        data_rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

        def ym_key(r):
            m = re.match(r'(\d{4})年(\d{2})月', str(r[0]))
            return int(m.group(1)) * 100 + int(m.group(2)) if m else 0

        data_rows.sort(key=ym_key)
        for i, row_vals in enumerate(data_rows):
            r = i + 2
            for col, val in enumerate(row_vals, 1):
                ws.cell(row=r, column=col, value=val)
            apply_row_style(ws, r)

        wb.save(history_path)
        return overwrite_decision

    # ------------------------------------------------------------------ 入金記録
    def _ensure_payment_sheet(self, wb):
        """wb に「入金履歴」シートがなければ作成する"""
        if "入金履歴" in wb.sheetnames:
            return
        ws2 = wb.create_sheet("入金履歴")
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        border   = _make_border()
        center   = Alignment(horizontal="center")
        for col, h in enumerate(["入金日", "入金額", "メモ"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = center; c.border = border
        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 28
        ws2.freeze_panes = "A2"

    def record_payment(self):
        """入金を記録するダイアログを開く"""
        if not self.settings.get("history_file"):
            if not self._choose_history_file():
                return

        import datetime
        dlg = tk.Toplevel(self.root)
        dlg.title("入金を記録")
        dlg.geometry("340x220")
        dlg.resizable(False, False)
        dlg.grab_set()

        pad = {"padx": 20, "pady": 4}
        today = datetime.date.today().strftime("%Y/%m/%d")

        tk.Label(dlg, text="入金日（例：2026/03/25）",
                 font=("", 9), fg="gray", anchor="w").pack(fill=tk.X, **pad)
        e_date = tk.Entry(dlg, font=("", 11))
        e_date.insert(0, today)
        e_date.pack(fill=tk.X, padx=20, pady=(0, 8))

        tk.Label(dlg, text="入金額（円）",
                 font=("", 9), fg="gray", anchor="w").pack(fill=tk.X, **pad)
        e_amount = tk.Entry(dlg, font=("", 11))
        e_amount.pack(fill=tk.X, padx=20, pady=(0, 8))

        tk.Label(dlg, text="メモ（任意）",
                 font=("", 9), fg="gray", anchor="w").pack(fill=tk.X, **pad)
        e_memo = tk.Entry(dlg, font=("", 11))
        e_memo.pack(fill=tk.X, padx=20, pady=(0, 8))

        def on_save():
            date_str = e_date.get().strip()
            try:
                amount = int(e_amount.get().replace(",", "").strip())
                if amount <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("エラー", "入金額に正しい金額を入力してください。", parent=dlg)
                return
            memo = e_memo.get().strip()
            try:
                self._save_payment(date_str, amount, memo)
                self.refresh_balance()
                messagebox.showinfo("完了", f"入金を記録しました。\n{date_str}　¥{amount:,}", parent=dlg)
                dlg.destroy()
            except Exception as ex:
                messagebox.showerror("エラー", f"記録に失敗しました。\n{ex}", parent=dlg)

        btn_f = tk.Frame(dlg)
        btn_f.pack(fill=tk.X, padx=20, pady=4)
        tk.Button(btn_f, text="記録する", command=on_save,
                  bg="#1f4e79", fg="white", font=("", 10, "bold"),
                  padx=12, pady=3, relief=tk.FLAT).pack(side=tk.RIGHT)
        tk.Button(btn_f, text="キャンセル", command=dlg.destroy,
                  padx=8, pady=3).pack(side=tk.RIGHT, padx=6)

    def _save_payment(self, date_str: str, amount: int, memo: str):
        """入金履歴シートに1行追加して保存"""
        history_path = self.settings["history_file"]
        if os.path.exists(history_path):
            wb = openpyxl.load_workbook(history_path)
        else:
            wb = openpyxl.Workbook()
            wb.active.title = "請求履歴"

        self._ensure_payment_sheet(wb)
        ws2    = wb["入金履歴"]
        border = _make_border()
        r      = ws2.max_row + 1
        cells  = [
            (1, date_str,  Alignment(horizontal="center")),
            (2, amount,    Alignment(horizontal="right")),
            (3, memo,      Alignment(horizontal="left")),
        ]
        for col, val, align in cells:
            c = ws2.cell(row=r, column=col, value=val)
            c.border = border
            c.font   = Font(size=10)
            c.alignment = align
            if col == 2:
                c.number_format = "#,##0"
        wb.save(history_path)

    def refresh_balance(self):
        """請求履歴・入金履歴を読み込んで残高サマリーを更新する"""
        history_path = self.settings.get("history_file", "")
        if not history_path or not os.path.exists(history_path):
            self.lbl_invoiced.config(text="累計請求：¥―")
            self.lbl_received.config(text="累計入金：¥―")
            self.lbl_balance.config(text="未収残高：¥―")
            return
        try:
            wb = openpyxl.load_workbook(history_path, read_only=True, data_only=True)

            # Sheet1：奥さん請求額（列3）の合計
            ws1 = wb.active
            total_invoiced = sum(
                (row[2].value or 0)
                for row in ws1.iter_rows(min_row=2)
                if row[2].value is not None
            )

            # Sheet2：入金額（列2）の合計
            total_received = 0
            if "入金履歴" in wb.sheetnames:
                ws2 = wb["入金履歴"]
                total_received = sum(
                    (row[1].value or 0)
                    for row in ws2.iter_rows(min_row=2)
                    if row[1].value is not None
                )
            wb.close()

            balance = total_invoiced - total_received
            color   = "#c00000" if balance > 0 else "#217346"

            self.lbl_invoiced.config(text=f"累計請求：¥{total_invoiced:,}")
            self.lbl_received.config(text=f"累計入金：¥{total_received:,}")
            self.lbl_balance.config(text=f"未収残高：¥{balance:,}", fg=color)
        except Exception:
            pass  # ファイルが開けない場合はサマリーを据え置き

    # ------------------------------------------------------------------ 設定ダイアログ
    def open_settings(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("設定")
        dlg.geometry("460x480")
        dlg.resizable(False, False)
        dlg.grab_set()

        # ---- 固定項目セクション
        tk.Label(dlg, text="固定請求項目", font=("", 11, "bold"), pady=8).pack()
        tk.Label(dlg, text="奥さんへの請求として自動チェックする店舗名（部分一致）",
                 font=("", 9), fg="gray").pack()

        frame = tk.Frame(dlg, padx=20)
        frame.pack(fill=tk.BOTH, expand=True, pady=4)

        listbox = tk.Listbox(frame, font=("", 10), height=8, selectmode=tk.SINGLE)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(frame, command=listbox.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.config(yscrollcommand=sb.set)
        for item in self.settings["fixed_items"]:
            listbox.insert(tk.END, item)

        entry_frame = tk.Frame(dlg, padx=20)
        entry_frame.pack(fill=tk.X, pady=4)
        entry = tk.Entry(entry_frame, font=("", 10))
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        def add_item():
            val = entry.get().strip()
            if val and val not in listbox.get(0, tk.END):
                listbox.insert(tk.END, val)
                entry.delete(0, tk.END)

        def delete_item():
            sel = listbox.curselection()
            if sel:
                listbox.delete(sel[0])

        tk.Button(entry_frame, text="追加", command=add_item,
                  bg="#217346", fg="white", padx=8).pack(side=tk.LEFT, padx=4)
        tk.Button(dlg, text="削除（選択中）", command=delete_item).pack(pady=2)

        # ---- 履歴ファイルセクション
        sep = tk.Frame(dlg, height=1, bg="#cccccc")
        sep.pack(fill=tk.X, padx=20, pady=8)

        tk.Label(dlg, text="履歴ファイルの保存先", font=("", 11, "bold")).pack()

        hist_frame = tk.Frame(dlg, padx=20)
        hist_frame.pack(fill=tk.X, pady=6)

        current_path = self.settings.get("history_file", "")
        display = os.path.basename(current_path) if current_path else "未設定"
        self.hist_path_label = tk.Label(hist_frame, text=display,
                                        fg="black" if current_path else "gray",
                                        font=("", 9), wraplength=340, justify="left")
        self.hist_path_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        def change_history_path():
            path = filedialog.asksaveasfilename(
                title="履歴ファイルの保存場所を選択",
                initialfile=os.path.basename(self.settings.get("history_file", "seikyu_history")),
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if path:
                self.settings["history_file"] = path
                self.hist_path_label.config(
                    text=os.path.basename(path), fg="black"
                )

        tk.Button(hist_frame, text="変更", command=change_history_path,
                  padx=8).pack(side=tk.RIGHT)

        # ---- 保存ボタン
        def save_and_close():
            self.settings["fixed_items"] = list(listbox.get(0, tk.END))
            save_settings(self.settings)
            # チェック状態を再計算（固定項目の変更を反映）
            if self.rows:
                children = self.tree.get_children()
                for idx, item_id in enumerate(children):
                    r = self.rows[idx]
                    if r["is_info"]:
                        continue
                    r["is_fixed"] = self.is_fixed(r["store"])
                    if r["is_fixed"]:
                        self.checked[idx] = True
                    values    = list(self.tree.item(item_id, "values"))
                    values[5] = "○" if self.checked[idx] else ""
                    tag = "fixed" if r["is_fixed"] else ("checked" if self.checked[idx] else "")
                    self.tree.item(item_id, values=values, tags=(tag,))
                self.update_total()
            dlg.destroy()

        tk.Button(dlg, text="保存して閉じる", command=save_and_close,
                  bg="#e60026", fg="white", font=("", 10, "bold"),
                  pady=5, padx=16).pack(pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    app = RakutenApp(root)
    root.mainloop()
